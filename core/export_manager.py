"""
Export manager for SQL metadata (Excel + JSON only).
"""
import json
import os
from abc import ABC, abstractmethod
from typing import Dict, List, Optional

import pandas as pd

from models.sql_metadata import SQLMetadata


class ExportStrategy(ABC):
    """Абстрактная стратегия экспорта метаданных SQL.

    Subclasses должны реализовать методы export и get_file_extensions.
    """
    @abstractmethod
    def export(self, metadata: SQLMetadata, file_path: str, **kwargs) -> bool:
        """Экспортирует метаданные в файл.

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к файлу для сохранения.
            **kwargs: Дополнительные параметры.

        Returns:
            True, если экспорт успешен, иначе False.
        """
        raise NotImplementedError

    @abstractmethod
    def get_file_extensions(self) -> List[str]:
        """Возвращает список поддерживаемых расширений файлов.

        Returns:
            Список строк, например [".xlsx", ".xls"].
        """
        raise NotImplementedError


class ExcelExportStrategy(ExportStrategy):
    """Стратегия экспорта в Excel (XLSX/XLS)."""

    def export(self, metadata: SQLMetadata, file_path: str, **kwargs) -> bool:
        """Экспортирует метаданные в Excel файл с несколькими листами.

        Создаёт листы:
            - Таблицы: список таблиц с алиасами и типами.
            - Колонки: детальная информация о колонках.
            - Линедж: зависимости и использование колонок.
            - Статистика: сводная статистика.
            - Текстовый вывод: текстовое представление результатов.
            - Исходный SQL: оригинальный SQL запрос.

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к файлу Excel.
            **kwargs: Дополнительные параметры (не используются).

        Returns:
            True, если экспорт успешен, иначе False.
        """
        try:
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            # Функция для фильтрации "calculation" из usage_locations
            def filter_calculation(locations):
                return [loc for loc in locations if loc.lower() != "calculation"]

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Лист "Колонки" (первый)
                columns_data = []
                for c in metadata.columns:
                    table = c.table or ""
                    schema = ""
                    table_name_without_schema = table
                    object_type = ""
                    # Попробуем найти таблицу в метаданных
                    if table:
                        # Разделяем схему и имя, если есть точка
                        if "." in table:
                            parts = table.split(".", 1)
                            schema = parts[0]
                            table_name_without_schema = parts[1]
                        # Ищем таблицу в метаданных (сначала по схеме, если нет - по имени)
                        found_table = None
                        if schema:
                            found_table = metadata.get_table_by_name(table_name_without_schema, schema)
                        if not found_table:
                            # Поиск по имени без схемы (первая подходящая)
                            for t in metadata.get_unique_tables():
                                if t.name == table_name_without_schema:
                                    found_table = t
                                    break
                        if found_table:
                            object_type = found_table.table_type.value
                            # Если схема не была извлечена из table, возьмём из found_table
                            if not schema:
                                schema = found_table.schema or ""
                    columns_data.append({
                        "Колонка": c.column_name or "",
                        "Полное имя колонки": c.full_name or "",
                        "Схема": schema,
                        "Таблица": table_name_without_schema,
                        "Алиас таблицы": c.table_alias or "",
                        "Тип объекта": object_type,
                        "Алиасы": c.get_aliases_str(),
                        "Где используется": ", ".join(filter_calculation(c.usage_locations)),
                        "Количество упоминаний": c.usage_count,
                    })
                columns_df = pd.DataFrame(columns_data)
                columns_df.to_excel(writer, sheet_name="Колонки", index=False)

                # Лист "Таблицы" (второй)
                tables_df = pd.DataFrame(
                    [
                        {
                            "Схема": t.schema or "",
                            "Таблица": t.name,
                            "Алиасы": t.get_aliases_str(),
                            "Тип": t.table_type.value,
                            "Колонок": len(t.columns),
                            "Тип JOIN": t.join_type or "",
                        }
                        for t in metadata.get_unique_tables()
                    ]
                )
                tables_df.to_excel(writer, sheet_name="Таблицы", index=False)

                # Лист "Линедж"
                lineage_df = pd.DataFrame(
                    [
                        {
                            "Колонка": c.column_name,
                            "Таблица": c.table or "",
                            "Источник": c.full_name or "",
                            "Зависимости": ", ".join(c.dependencies),
                            "Использование": ", ".join(filter_calculation(c.usage_locations)),
                        }
                        for c in metadata.columns
                    ]
                )
                lineage_df.to_excel(writer, sheet_name="Линедж", index=False)

                # Лист "Статистика" (таблица параметров)
                stats = metadata.get_statistics()
                stats_rows = []
                for key, value in stats.items():
                    if isinstance(value, dict):
                        for subkey, subvalue in value.items():
                            stats_rows.append({"Параметр": f"{key}.{subkey}", "Значение": subvalue})
                    else:
                        stats_rows.append({"Параметр": key, "Значение": value})
                stats_df = pd.DataFrame(stats_rows)
                stats_df.to_excel(writer, sheet_name="Статистика", index=False)

                # Функция для экранирования текста, который Excel может интерпретировать как формулу
                def escape_excel_text(text: str) -> str:
                    if not isinstance(text, str):
                        return text
                    # Если строка начинается с =, -, +, @, добавляем обратный апостроф
                    if text.startswith(('=', '-', '+', '@')):
                        return '`' + text
                    return text

                # Лист "Текстовый вывод"
                text_lines = []
                text_lines.append("=== РЕЗУЛЬТАТЫ ПАРСИНГА SQL ===")
                text_lines.append(f"Дата анализа: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                text_lines.append(f"Диалект SQL: {metadata.dialect if hasattr(metadata, 'dialect') else 'Не указан'}")
                text_lines.append("")
                text_lines.append("=== УНИКАЛЬНЫЕ ТАБЛИЦЫ ===")
                for i, table in enumerate(metadata.get_unique_tables(), 1):
                    aliases_info = f" (алиасы: {table.get_aliases_str()})" if table.aliases else ""
                    join_info = f" [{table.join_type}]" if table.join_type else ""
                    text_lines.append(f"{i}. {table.schema or ''}.{table.name}{aliases_info} - {table.table_type.value}{join_info}")
                text_lines.append(f"Всего уникальных таблиц: {len(metadata.get_unique_tables())}")
                text_lines.append("")
                text_lines.append("=== ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ ===")
                for i, col in enumerate(metadata.column_analysis, 1):
                    text_lines.append(f"{i}. Колонка: {col.column_name}")
                    text_lines.append(f"   Таблица: {col.table or ''} (Алиас: {col.table_alias or ''})")
                    text_lines.append(f"   Полное имя: {col.full_name or ''}")
                    text_lines.append(f"   Алиасы: {col.get_aliases_str()}")
                    text_lines.append(f"   Использование: {', '.join(filter_calculation(col.usage_locations))}")
                    text_lines.append(f"   Упоминаний: {col.usage_count}")
                    text_lines.append("")
                # Экранируем строки
                escaped_text_lines = [escape_excel_text(line) for line in text_lines]
                text_df = pd.DataFrame({"Текст": escaped_text_lines})
                text_df.to_excel(writer, sheet_name="Текстовый вывод", index=False)

                # Лист "Исходный SQL"
                # Экранируем SQL: добавляем обратный апостроф и убираем ведущий перевод строки
                sql_text = metadata.original_sql
                if isinstance(sql_text, str):
                    # Убираем ведущие переводы строки
                    sql_text = sql_text.lstrip('\n')
                    # Добавляем обратный апостроф, если ещё нет
                    if not sql_text.startswith('`'):
                        sql_text = '`' + sql_text
                sql_df = pd.DataFrame([{"Исходный SQL": sql_text}])
                sql_df.to_excel(writer, sheet_name="Исходный SQL", index=False)

                # Автоширина колонок и автофильтр
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    ws.auto_filter.ref = ws.dimensions
                    # Постобработка не требуется, так как экранирование выполнено ранее
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        max_len = max(len(str(cell.value or "")) for cell in col)
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            return True
        except Exception:
            return False

    def get_file_extensions(self) -> List[str]:
        """Возвращает поддерживаемые расширения файлов Excel.

        Returns:
            Список [".xlsx", ".xls"].
        """
        return [".xlsx", ".xls"]


class JSONExportStrategy(ExportStrategy):
    """Стратегия экспорта в JSON."""

    def export(self, metadata: SQLMetadata, file_path: str, **kwargs) -> bool:
        """Экспортирует метаданные в JSON файл.

        Структура JSON:
            {
                "metadata": {
                    "original_sql": ...,
                    "statistics": ...,
                    "parse_errors": ...
                },
                "tables": [...],
                "columns": [...],
                "json_schema": {...}
            }

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к файлу JSON.
            **kwargs: Дополнительные параметры (не используются).

        Returns:
            True, если экспорт успешен, иначе False.
        """
        try:
            payload = {
                "metadata": {
                    "original_sql": metadata.original_sql,
                    "statistics": metadata.get_statistics(),
                    "parse_errors": metadata.parse_errors,
                },
                "tables": [table.to_dict() for table in metadata.get_unique_tables()],
                "columns": [column.to_dict() for column in metadata.columns],
                "json_schema": metadata.json_schema,
            }
            with open(file_path, "w", encoding="utf-8") as fp:
                json.dump(payload, fp, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False

    def get_file_extensions(self) -> List[str]:
        """Возвращает поддерживаемые расширения файлов JSON.

        Returns:
            Список [".json"].
        """
        return [".json"]


class CSVExportStrategy(ExportStrategy):
    """Стратегия экспорта в CSV (таблицы и колонки в отдельные файлы)."""

    def export(self, metadata: SQLMetadata, file_path: str, **kwargs) -> bool:
        """Экспортирует метаданные в CSV файлы.

        Создаёт два файла:
            - <basename>_tables.csv: таблицы
            - <basename>_columns.csv: колонки

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к основному файлу CSV (будет использован для генерации имён).
            **kwargs: Дополнительные параметры (не используются).

        Returns:
            True, если экспорт успешен, иначе False.
        """
        try:
            import pandas as pd
            import os

            base, _ = os.path.splitext(file_path)
            tables_path = f"{base}_tables.csv"
            columns_path = f"{base}_columns.csv"

            # Таблицы
            tables_data = []
            for table in metadata.get_unique_tables():
                tables_data.append({
                    "Схема": table.schema or "",
                    "Имя таблицы": table.name,
                    "Алиасы": table.get_aliases_str(),
                    "Тип таблицы": table.table_type.value,
                    "Тип JOIN": table.join_type or "",
                    "Колонки": ", ".join(table.columns) if table.columns else ""
                })
            if tables_data:
                df_tables = pd.DataFrame(tables_data)
                df_tables.to_csv(tables_path, index=False, encoding='utf-8-sig')
            else:
                # Создать пустой файл
                with open(tables_path, 'w', encoding='utf-8-sig') as f:
                    f.write("Схема,Имя таблицы,Алиасы,Тип таблицы,Тип JOIN,Колонки\n")

            # Колонки
            columns_data = []
            for col in metadata.columns:
                table = col.table or ""
                schema = ""
                table_name_without_schema = table
                object_type = ""
                if table:
                    if "." in table:
                        parts = table.split(".", 1)
                        schema = parts[0]
                        table_name_without_schema = parts[1]
                    # Ищем таблицу в метаданных
                    found_table = None
                    if schema:
                        found_table = metadata.get_table_by_name(table_name_without_schema, schema)
                    if not found_table:
                        # Поиск по имени без схемы (первая подходящая)
                        for t in metadata.get_unique_tables():
                            if t.name == table_name_without_schema:
                                found_table = t
                                break
                    if found_table:
                        object_type = found_table.table_type.value
                        # Если схема не была извлечена из table, возьмём из found_table
                        if not schema:
                            schema = found_table.schema or ""
                columns_data.append({
                    "Колонка": col.column_name,
                    "Полное имя": col.full_name or "",
                    "Схема": schema,
                    "Таблица": table_name_without_schema,
                    "Алиас таблицы": col.table_alias or "",
                    "Тип объекта": object_type,
                    "Тип вычисления": col.calculation_type or "",
                    "Использование": ", ".join(col.usage_locations) if col.usage_locations else "",
                    "Количество использований": col.usage_count
                })
            if columns_data:
                df_columns = pd.DataFrame(columns_data)
                df_columns.to_csv(columns_path, index=False, encoding='utf-8-sig')
            else:
                with open(columns_path, 'w', encoding='utf-8-sig') as f:
                    f.write("Колонка,Полное имя,Схема,Таблица,Алиас таблицы,Тип объекта,Тип вычисления,Использование,Количество использований\n")

            return True
        except Exception:
            return False

    def get_file_extensions(self) -> List[str]:
        """Возвращает поддерживаемые расширения файлов CSV.

        Returns:
            Список [".csv"].
        """
        return [".csv"]


class TextExportStrategy(ExportStrategy):
    """Стратегия экспорта в текстовый файл (plain text)."""

    def export(self, metadata: SQLMetadata, file_path: str, **kwargs) -> bool:
        """Экспортирует метаданные в текстовый файл.

        Формат аналогичен вкладке "Текстовый вывод".

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к файлу.
            **kwargs: Дополнительные параметры (не используются).

        Returns:
            True, если экспорт успешен, иначе False.
        """
        try:
            from datetime import datetime
            text = f"=== РЕЗУЛЬТАТЫ ПАРСИНГА SQL ===\n"
            text += f"Дата анализа: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            text += f"Диалект SQL: {kwargs.get('dialect', 'ORACLE')}\n\n"

            # Уникальные таблицы
            text += "=== УНИКАЛЬНЫЕ ТАБЛИЦЫ ===\n"
            for i, table in enumerate(metadata.get_unique_tables(), 1):
                aliases_info = f" (алиасы: {table.get_aliases_str()})" if table.aliases else ""
                join_info = f" [{table.join_type}]" if table.join_type else ""
                text += f"{i}. {table.schema or ''}.{table.name}{aliases_info} - {table.table_type.value}{join_info}\n"
            text += f"\nВсего уникальных таблиц: {len(metadata.get_unique_tables())}\n\n"

            # Детальная информация о колонках
            text += "=== ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ ===\n\n"
            for i, col in enumerate(metadata.columns, 1):
                text += f"{i}. Колонка: {col.column_name}\n"
                text += f"   Таблица: {col.table or ''} "
                text += f"(Алиас: {col.table_alias or ''})\n"
                text += f"   Полное имя: {col.full_name or ''}\n"
                text += f"   Вычисление: {col.calculation_type or 'нет'}\n"
                filtered_locations = [loc for loc in col.usage_locations if loc.lower() != "calculation"]
                text += f"   Использование: {', '.join(filtered_locations) if filtered_locations else 'нет'}\n"
                text += f"{'-'*40}\n"

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(text)
            return True
        except Exception:
            return False

    def get_file_extensions(self) -> List[str]:
        """Возвращает поддерживаемые расширения файлов текста.

        Returns:
            Список [".txt"].
        """
        return [".txt"]


class ExportManager:
    """Менеджер экспорта, который выбирает стратегию по расширению файла или имени."""

    def __init__(self):
        """Инициализирует менеджер с доступными стратегиями экспорта."""
        self._strategies: Dict[str, ExportStrategy] = {
            "excel": ExcelExportStrategy(),
            "json": JSONExportStrategy(),
            "csv": CSVExportStrategy(),
            "text": TextExportStrategy(),
        }

    def export(self, metadata: SQLMetadata, file_path: str, strategy_name: Optional[str] = None, **kwargs) -> bool:
        """Экспортирует метаданные, используя указанную стратегию или определяя по расширению файла.

        Args:
            metadata: Объект SQLMetadata.
            file_path: Путь к файлу для сохранения.
            strategy_name: Имя стратегии ("excel", "json", "csv", "text"). Если None, определяется по расширению файла.
            **kwargs: Дополнительные параметры, передаваемые стратегии.

        Returns:
            True, если экспорт успешен, иначе False.

        Raises:
            ValueError: Если не удалось определить стратегию экспорта.
        """
        strategy = self._strategies.get(strategy_name) if strategy_name else None
        if strategy is None:
            ext = os.path.splitext(file_path)[1].lower()
            for candidate in self._strategies.values():
                if any(ext == supported for supported in candidate.get_file_extensions()):
                    strategy = candidate
                    break
        if strategy is None:
            raise ValueError(f"Не удалось определить стратегию экспорта: {file_path}")
        return strategy.export(metadata, file_path, **kwargs)

    def get_available_strategies(self) -> Dict[str, ExportStrategy]:
        """Возвращает словарь доступных стратегий экспорта.

        Returns:
            Словарь {имя_стратегии: объект_стратегии}.
        """
        return self._strategies.copy()
